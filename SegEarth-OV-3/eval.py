import os
import os.path as osp
import argparse
import csv
import json
import openpyxl
from mmengine.runner import Runner
from mmengine.config import Config, DictAction

import segearthov3_segmentor
import custom_datasets


def parse_args():
    parser = argparse.ArgumentParser(
        description='CorrCLIP evaluation with MMSeg')
    parser.add_argument('config', default='./configs/cfg_loveda.py')
    parser.add_argument(
        '--show', action='store_true', help='show prediction results')
    parser.add_argument(
        '--show_dir',
        default='./show_dir/',
        help='directory to save visualizaion images')
    parser.add_argument(
        '--out',
        type=str,
        help='The directory to save output prediction for offline evaluation')
    parser.add_argument(
        '--cfg-options',
        nargs='+',
        action=DictAction,
        help='override some settings in the used config, the key-value pair '
        'in xxx=yyy format will be merged into config file. If the value to '
        'be overwritten is a list, it should be like key="[a,b]" or key=a,b '
        'It also allows nested list/tuple values, e.g. key="[(a,b),(c,d)]" '
        'Note that the quotation marks are necessary and that no white space '
        'is allowed.')
    parser.add_argument(
        '--launcher',
        choices=['none', 'pytorch', 'slurm', 'mpi'],
        default='none',
        help='job launcher')
    # When using PyTorch version >= 2.0.0, the `torch.distributed.launch`
    # will pass the `--local-rank` parameter to `tools/train.py` instead
    # of `--local_rank`.
    parser.add_argument('--local_rank', '--local-rank', type=int, default=0)
    args = parser.parse_args()
    if 'LOCAL_RANK' not in os.environ:
        os.environ['LOCAL_RANK'] = str(args.local_rank)

    return args


def append_experiment_result(file_path, experiment_data):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    if sheet['A1'].value is None:
        sheet['A1'] = 'Model'
        sheet['B1'] = 'Dataset'
        sheet['C1'] = 'aAcc'
        sheet['D1'] = 'mIoU'
        sheet['E1'] = 'mAcc'

    last_row = sheet.max_row

    for index, result in enumerate(experiment_data, start=1):
        sheet.cell(row=last_row + index, column=1, value=result['Model'])
        sheet.cell(row=last_row + index, column=2, value=result['Dataset'])
        sheet.cell(row=last_row + index, column=3, value=result['aAcc'])
        sheet.cell(row=last_row + index, column=4, value=result['mIoU'])
        sheet.cell(row=last_row + index, column=5, value=result['mAcc'])

    workbook.save(file_path)


def save_classwise_results(runner, output_dir):
    if not hasattr(runner, 'test_evaluator'):
        return None

    metrics = getattr(runner.test_evaluator, 'metrics', None)
    if not metrics:
        return None

    metric = metrics[0]
    if not hasattr(metric, 'results') or not metric.results:
        return None

    if not hasattr(metric, 'total_area_to_metrics'):
        return None

    results = tuple(zip(*metric.results))
    if len(results) != 4:
        return None

    total_area_intersect = sum(results[0])
    total_area_union = sum(results[1])
    total_area_pred_label = sum(results[2])
    total_area_label = sum(results[3])

    ret_metrics = metric.total_area_to_metrics(
        total_area_intersect,
        total_area_union,
        total_area_pred_label,
        total_area_label,
        metric.metrics,
        getattr(metric, 'nan_to_num', 0),
        getattr(metric, 'beta', 1.0),
    )

    class_names = metric.dataset_meta['classes']
    a_acc = ret_metrics.pop('aAcc', None)
    class_iou = [round(float(value) * 100, 2) for value in ret_metrics['IoU']]
    class_acc = [round(float(value) * 100, 2) for value in ret_metrics['Acc']]
    rows = [
        {'Class': class_name, 'IoU': iou_value, 'Acc': acc_value}
        for class_name, iou_value, acc_value in zip(class_names, class_iou, class_acc)
    ]

    os.makedirs(output_dir, exist_ok=True)
    csv_path = os.path.join(output_dir, 'class_iou_results.csv')
    json_path = os.path.join(output_dir, 'class_iou_results.json')
    txt_path = os.path.join(output_dir, 'class_iou_results.txt')

    with open(csv_path, 'w', newline='') as handle:
        writer = csv.DictWriter(handle, fieldnames=['Class', 'IoU', 'Acc'])
        writer.writeheader()
        writer.writerows(rows)

    summary = {
        'aAcc': round(float(a_acc) * 100, 2) if a_acc is not None else None,
        'mIoU': round(float(ret_metrics['IoU'].mean()) * 100, 2),
        'mAcc': round(float(ret_metrics['Acc'].mean()) * 100, 2),
        'classes': rows,
    }
    with open(json_path, 'w') as handle:
        json.dump(summary, handle, ensure_ascii=False, indent=2)

    with open(txt_path, 'w') as handle:
        handle.write('Class\tIoU\tAcc\n')
        for row in rows:
            handle.write(f"{row['Class']}\t{row['IoU']}\t{row['Acc']}\n")
        handle.write(f"\nmIoU\t{summary['mIoU']}\n")
        handle.write(f"mAcc\t{summary['mAcc']}\n")

    return {
        'csv_path': csv_path,
        'json_path': json_path,
        'txt_path': txt_path,
    }


def trigger_visualization_hook(cfg, args):
    default_hooks = cfg.default_hooks
    if 'visualization' in default_hooks:
        visualization_hook = default_hooks['visualization']
        # Turn on visualization
        visualization_hook['draw'] = True
        if args.show:
            visualization_hook['show'] = True
            visualization_hook['wait_time'] = args.wait_time
        if args.show_dir:
            visualizer = cfg.visualizer
            visualizer['save_dir'] = args.show_dir
    else:
        raise RuntimeError(
            'VisualizationHook must be included in default_hooks.'
            'refer to usage '
            '"visualization=dict(type=\'VisualizationHook\')"')

    return cfg


def main():
    args = parse_args()
    print(os.getcwd())
    cfg = Config.fromfile(args.config)
    cfg.launcher = args.launcher
    # add output_dir in metric
    if args.out is not None:
        cfg.test_evaluator['output_dir'] = args.out
        cfg.test_evaluator['keep_results'] = True
    if args.cfg_options is not None:
        cfg.merge_from_dict(args.cfg_options)
    cfg.work_dir = osp.join('./work_dirs',
                            osp.splitext(osp.basename(args.config))[0])

    # trigger_visualization_hook(cfg, args)
    runner = Runner.from_cfg(cfg)
    results = runner.test()

    results.update({'Model': cfg.model.model_type,
                    'Dataset': cfg.dataset_type})

    if runner.rank == 0:
        append_experiment_result('results.xlsx', [results])

    if runner.rank == 0:
        with open(os.path.join(cfg.work_dir, 'results.txt'), 'a') as f:
            f.write(os.path.basename(args.config).split('.')[0] + '\n')
            for k, v in results.items():
                f.write(k + ': ' + str(v) + '\n')

        classwise_paths = save_classwise_results(runner, args.out or cfg.work_dir)
        if classwise_paths is not None:
            with open(os.path.join(cfg.work_dir, 'results.txt'), 'a') as f:
                f.write('Classwise results saved to:\n')
                for key, value in classwise_paths.items():
                    f.write(f'{key}: {value}\n')


if __name__ == '__main__':
    main()
