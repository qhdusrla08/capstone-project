"""
eval_with_adapter_fusion.py — SegEarth-OV-3 + RSAdapter + fusion mode evaluation

eval_with_adapter.py에 test_fusion_modes.py의 head fusion 전략을 추가한 평가 스크립트.
RSAdapter를 SAM3 ViT 블록에 forward hook으로 등록한 뒤,
--fusion_mode 인자로 fusion 전략을 선택할 수 있다:
  max            : element-wise max (기본, baseline)
  heuristic      : things/stuff에 따라 α 가중합 (--things_alpha / --stuff_alpha)
  entropy        : per-pixel binary entropy 기반 adaptive weighting
  inst_only      : instance head만 사용 (semantic head 무시)
  sem_only       : semantic head만 사용 (instance head 무시)
  adaptive_split : things=inst only(α=1.0), stuff=sem only(α=0.0)

실행 예시:
  cd ~/capstone/SegEarth-OV-3

  # RSAdapter + max (baseline)
  python eval_with_adapter_fusion.py configs/cfg_loveda.py \
      --adapter_ckpt rs_adapter/ckpt_full_best.pt \
      --fusion_mode max

  # RSAdapter + heuristic
  python eval_with_adapter_fusion.py configs/cfg_loveda.py \
      --adapter_ckpt rs_adapter/ckpt_full_best.pt \
      --fusion_mode heuristic --things_alpha 0.8 --stuff_alpha 0.2 \
      2>&1 | tee eval_adapter_fusion_heuristic.txt

  # RSAdapter + entropy
  python eval_with_adapter_fusion.py configs/cfg_loveda.py \
      --adapter_ckpt rs_adapter/ckpt_full_best.pt \
      --fusion_mode entropy \
      2>&1 | tee eval_adapter_fusion_entropy.txt
"""

import os
import os.path as osp
import argparse
import types
import numpy as np
import openpyxl
import torch
import torch.nn as nn
import torch.nn.functional as F
from mmengine.runner import Runner
from mmengine.config import Config, DictAction

import segearthov3_segmentor
import custom_datasets

sys_path_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rs_adapter")
import sys
sys.path.insert(0, sys_path_dir)
from rs_adapter import RSAdapter

FUSION_MODES = ["max", "heuristic", "entropy", "inst_only", "sem_only", "adaptive_split"]


def parse_args():
    parser = argparse.ArgumentParser(
        description='SegEarth-OV-3 + RSAdapter + fusion mode evaluation')
    parser.add_argument('config', default='./configs/cfg_loveda.py')
    parser.add_argument(
        '--adapter_ckpt', required=True,
        help='RSAdapter 체크포인트 경로 (ckpt_full_best.pt)')
    parser.add_argument(
        '--fusion_mode', choices=FUSION_MODES, default='max',
        help='head fusion 전략 선택')
    parser.add_argument(
        '--things_alpha', type=float, default=0.8,
        help='heuristic 모드: things 클래스에 적용할 instance head 가중치 (default: 0.8)')
    parser.add_argument(
        '--stuff_alpha', type=float, default=0.2,
        help='heuristic 모드: stuff 클래스에 적용할 instance head 가중치 (default: 0.2)')
    parser.add_argument(
        '--things_set', nargs='+', default=['building', 'house', 'roof'],
        help='things(개체성) 클래스로 분류할 클래스명 목록 (default: building house roof)')
    parser.add_argument('--show', action='store_true')
    parser.add_argument('--show_dir', default='./show_dir/')
    parser.add_argument('--out', type=str)
    parser.add_argument(
        '--cfg-options', nargs='+', action=DictAction,
        help='xxx=yyy 형태로 config를 override')
    parser.add_argument(
        '--launcher', choices=['none', 'pytorch', 'slurm', 'mpi'], default='none')
    parser.add_argument('--local_rank', '--local-rank', type=int, default=0)
    args = parser.parse_args()
    if 'LOCAL_RANK' not in os.environ:
        os.environ['LOCAL_RANK'] = str(args.local_rank)
    return args


def append_experiment_result(file_path, experiment_data):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    sheet = workbook.active
    if sheet['A1'].value is None:
        sheet['A1'] = 'Model'
        sheet['B1'] = 'Dataset'
        sheet['C1'] = 'aAcc'
        sheet['D1'] = 'mIoU'
        sheet['E1'] = 'mAcc'
    last_row = sheet.max_row
    for index, result in enumerate(experiment_data, start=1):
        sheet.cell(row=last_row + index, column=1, value=result['Model'])
        sheet.cell(row=last_row + index, column=2, value=result['Dataset'])
        sheet.cell(row=last_row + index, column=3, value=result['aAcc'])
        sheet.cell(row=last_row + index, column=4, value=result['mIoU'])
        sheet.cell(row=last_row + index, column=5, value=result['mAcc'])
    workbook.save(file_path)


def make_fused_inference(fusion_mode, things_alpha, stuff_alpha, things_set):
    """monkey-patch용 _inference_single_view를 반환한다."""
    things_set = set(things_set)
    EPS = 1e-6

    def _binary_entropy(p):
        p = p.clamp(EPS, 1.0 - EPS)
        return -(p * p.log() + (1 - p) * (1 - p).log()) / np.log(2)  # [0,1]

    def _inference_single_view(self, image):
        w, h = image.size
        seg_logits = torch.zeros((self.num_queries, h, w), device=self.device)

        with torch.no_grad(), torch.autocast(device_type="cuda", dtype=torch.bfloat16):
            inference_state = self.processor.set_image(image)

            for query_idx, query_word in enumerate(self.query_words):
                self.processor.reset_all_prompts(inference_state)
                inference_state = self.processor.set_text_prompt(
                    state=inference_state, prompt=query_word)

                # ── Instance head: MAX_i(P_inst_i × score_i) ─────────────
                if self.use_transformer_decoder:
                    if inference_state['masks_logits'].shape[0] > 0:
                        for inst_id in range(inference_state['masks_logits'].shape[0]):
                            instance_logits = inference_state['masks_logits'][inst_id].squeeze()
                            instance_score  = inference_state['object_score'][inst_id]
                            if instance_logits.shape != (h, w):
                                instance_logits = F.interpolate(
                                    instance_logits.view(1, 1, *instance_logits.shape),
                                    size=(h, w), mode='bilinear', align_corners=False,
                                ).squeeze()
                            seg_logits[query_idx] = torch.max(
                                seg_logits[query_idx], instance_logits * instance_score
                            )

                # ── Semantic head + Fusion ────────────────────────────────
                if self.use_sem_seg:
                    sem_logit = inference_state['semantic_mask_logits']
                    if sem_logit.shape != (h, w):
                        sem_logit = F.interpolate(
                            sem_logit if sem_logit.dim() == 4
                            else sem_logit.unsqueeze(0).unsqueeze(0),
                            size=(h, w), mode='bilinear', align_corners=False,
                        ).squeeze()

                    inst_agg = seg_logits[query_idx].clone()

                    if fusion_mode == "max":
                        seg_logits[query_idx] = torch.max(inst_agg, sem_logit)

                    elif fusion_mode == "entropy":
                        p_inst = torch.sigmoid(inst_agg.float())
                        p_sem  = torch.sigmoid(sem_logit.float())
                        h_inst = _binary_entropy(p_inst)
                        h_sem  = _binary_entropy(p_sem)
                        conf_inst = 1.0 - h_inst
                        conf_sem  = 1.0 - h_sem
                        total     = (conf_inst + conf_sem).clamp(min=EPS)
                        alpha_map = conf_inst / total  # per-pixel weight for inst head
                        seg_logits[query_idx] = (
                            alpha_map * inst_agg.float()
                            + (1.0 - alpha_map) * sem_logit.float()
                        )

                    elif fusion_mode == "inst_only":
                        pass  # inst_agg 그대로 유지, sem_logit 무시

                    elif fusion_mode == "sem_only":
                        seg_logits[query_idx] = sem_logit.float()

                    else:  # heuristic / adaptive_split
                        cls_name = query_word.split(",")[0].strip().lower()
                        if fusion_mode == "adaptive_split":
                            alpha = 1.0 if cls_name in things_set else 0.0
                        else:  # heuristic
                            alpha = things_alpha if cls_name in things_set else stuff_alpha
                        seg_logits[query_idx] = (
                            alpha * inst_agg.float()
                            + (1.0 - alpha) * sem_logit.float()
                        )

                # ── Presence score ────────────────────────────────────────
                if self.use_presence_score:
                    seg_logits[query_idx] = (
                        seg_logits[query_idx] * inference_state["presence_score"]
                    )

        return seg_logits

    return _inference_single_view


def main():
    args = parse_args()
    print(os.getcwd())

    # ── 1. Runner 생성 ──────────────────────────────────────────────────────
    cfg = Config.fromfile(args.config)
    cfg.launcher = args.launcher
    if args.out is not None:
        cfg.test_evaluator['output_dir'] = args.out
        cfg.test_evaluator['keep_results'] = True
    if args.cfg_options is not None:
        cfg.merge_from_dict(args.cfg_options)
    cfg.work_dir = osp.join('./work_dirs', osp.splitext(osp.basename(args.config))[0])

    runner = Runner.from_cfg(cfg)

    # ── 2. SAM3 ViT 블록 접근 ──────────────────────────────────────────────
    seg_model = runner.model
    sam3_model = seg_model.processor.model
    vit_blocks = sam3_model.backbone.vision_backbone.trunk.blocks
    num_blocks = len(vit_blocks)  # 32

    # ── 3. 체크포인트에서 설정 읽기 ────────────────────────────────────────
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    ckpt = torch.load(args.adapter_ckpt, map_location=device)
    bottleneck  = ckpt.get("config", {}).get("bottleneck", 64)
    in_channels = ckpt.get("config", {}).get("in_channels", 1024)
    print(f"[Adapter] RSAdapter 설정: d_model={in_channels}, bottleneck={bottleneck}")

    # ── 4. Adapter 초기화 및 가중치 로드 ───────────────────────────────────
    adapters = nn.ModuleList([
        RSAdapter(d_model=in_channels, bottleneck=bottleneck)
        for _ in range(num_blocks)
    ]).to(device)
    adapters.load_state_dict(ckpt["adapters"])
    adapters.eval()
    print(f"[Adapter] 가중치 로드 완료: {args.adapter_ckpt}")

    # ── 5. ViT 블록에 forward hook 등록 ────────────────────────────────────
    hooks = []
    for i in range(num_blocks):
        adapter = adapters[i]

        def _hook(_module, _input, output, _adapter=adapter):
            return _adapter(output)

        hooks.append(vit_blocks[i].register_forward_hook(_hook))
    print(f"[Adapter] Forward hook {len(hooks)}개 등록 완료 (ViT 블록 0~{num_blocks-1})")

    # ── 6. Fusion mode 주입 (monkey-patch) ─────────────────────────────────
    fused_fn = make_fused_inference(
        args.fusion_mode, args.things_alpha, args.stuff_alpha, args.things_set
    )
    runner.model._inference_single_view = types.MethodType(fused_fn, runner.model)
    print(f"[Fusion] mode={args.fusion_mode}  |  "
          f"things_alpha={args.things_alpha}  |  stuff_alpha={args.stuff_alpha}  |  "
          f"things_set={args.things_set}")

    # ── 7. 평가 실행 ────────────────────────────────────────────────────────
    results = runner.test()

    # ── 8. Hook 해제 ────────────────────────────────────────────────────────
    for h in hooks:
        h.remove()
    print("[Adapter] Forward hook 해제 완료")

    # ── 9. 결과 저장 ────────────────────────────────────────────────────────
    model_tag = cfg.model.model_type + f'+RSAdapter+fusion_{args.fusion_mode}'
    if args.fusion_mode == 'heuristic':
        model_tag += f'(t={args.things_alpha},s={args.stuff_alpha})'
    results.update({'Model': model_tag, 'Dataset': cfg.dataset_type})

    if runner.rank == 0:
        append_experiment_result('results.xlsx', [results])

    if runner.rank == 0:
        tag = (osp.basename(args.config).split('.')[0]
               + f'+adapter+fusion_{args.fusion_mode}')
        with open(osp.join(cfg.work_dir, 'results.txt'), 'a') as f:
            f.write(tag + '\n')
            for k, v in results.items():
                f.write(k + ': ' + str(v) + '\n')


if __name__ == '__main__':
    main()
