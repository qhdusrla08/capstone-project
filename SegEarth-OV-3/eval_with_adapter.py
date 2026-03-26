"""
eval_with_adapter.py — Phase 3 OVSS 평가 스크립트

Phase 2(eval.py)와 동일한 mmengine Runner 기반 OVSS 추론을 수행하되,
ckpt_full_best.pt의 RSAdapter를 SAM3 ViT 블록에 forward hook으로 등록한다.

실행 예시:
  cd ~/capstone/SegEarth-OV-3
  python eval_with_adapter.py configs/cfg_loveda.py \
      --adapter_ckpt rs_adapter/ckpt_full_best.pt \
      --cfg-options \
      test_dataloader.dataset.data_root=~/datasets/LoveDA_mmseg \
      2>&1 | tee eval_phase3_log.txt
"""

import os
import os.path as osp
import argparse
import openpyxl
import torch
import torch.nn as nn
from mmengine.runner import Runner
from mmengine.config import Config, DictAction

import segearthov3_segmentor
import custom_datasets

sys_path_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rs_adapter")

import sys
sys.path.insert(0, sys_path_dir)
from rs_adapter import RSAdapter


def parse_args():
    parser = argparse.ArgumentParser(
        description='SegEarth-OV-3 + RSAdapter OVSS evaluation (Phase 3)')
    parser.add_argument('config', default='./configs/cfg_loveda.py')
    parser.add_argument(
        '--adapter_ckpt',
        required=True,
        help='RSAdapter 체크포인트 경로 (ckpt_full_best.pt)')
    parser.add_argument(
        '--show', action='store_true', help='show prediction results')
    parser.add_argument(
        '--show_dir',
        default='./show_dir/',
        help='directory to save visualization images')
    parser.add_argument(
        '--out',
        type=str,
        help='The directory to save output prediction for offline evaluation')
    parser.add_argument(
        '--cfg-options',
        nargs='+',
        action=DictAction,
        help='override some settings in the used config, the key-value pair '
        'in xxx=yyy format will be merged into config file.')
    parser.add_argument(
        '--launcher',
        choices=['none', 'pytorch', 'slurm', 'mpi'],
        default='none',
        help='job launcher')
    parser.add_argument('--local_rank', '--local-rank', type=int, default=0)
    args = parser.parse_args()
    if 'LOCAL_RANK' not in os.environ:
        os.environ['LOCAL_RANK'] = str(args.local_rank)
    return args


def append_experiment_result(file_path, experiment_data):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    if sheet['A1'].value is None:
        sheet['A1'] = 'Model'
        sheet['B1'] = 'Dataset'
        sheet['C1'] = 'aAcc'
        sheet['D1'] = 'mIoU'
        sheet['E1'] = 'mAcc'

    last_row = sheet.max_row

    for index, result in enumerate(experiment_data, start=1):
        sheet.cell(row=last_row + index, column=1, value=result['Model'])
        sheet.cell(row=last_row + index, column=2, value=result['Dataset'])
        sheet.cell(row=last_row + index, column=3, value=result['aAcc'])
        sheet.cell(row=last_row + index, column=4, value=result['mIoU'])
        sheet.cell(row=last_row + index, column=5, value=result['mAcc'])

    workbook.save(file_path)


def main():
    args = parse_args()
    print(os.getcwd())

    # ── 1. mmengine Runner 생성 ─────────────────────────────────────────────
    cfg = Config.fromfile(args.config)
    cfg.launcher = args.launcher
    if args.out is not None:
        cfg.test_evaluator['output_dir'] = args.out
        cfg.test_evaluator['keep_results'] = True
    if args.cfg_options is not None:
        cfg.merge_from_dict(args.cfg_options)
    cfg.work_dir = osp.join('./work_dirs',
                            osp.splitext(osp.basename(args.config))[0])

    runner = Runner.from_cfg(cfg)

    # ── 2. SAM3 ViT 블록 접근 ──────────────────────────────────────────────
    # runner.model → SegEarthOV3Segmentation
    # .processor   → Sam3Processor
    # .model       → SAM3 image model
    # .backbone.vision_backbone.trunk.blocks → ViT blocks (nn.ModuleList, len=32)
    seg_model = runner.model
    sam3_model = seg_model.processor.model
    vit_blocks = sam3_model.backbone.vision_backbone.trunk.blocks
    num_blocks = len(vit_blocks)  # 32

    # ── 3. 체크포인트에서 bottleneck 크기 읽기 ─────────────────────────────
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    ckpt = torch.load(args.adapter_ckpt, map_location=device)

    # bottleneck은 ckpt["config"]에 저장되어 있음 (train_adapter.py 기준)
    bottleneck = ckpt.get("config", {}).get("bottleneck", 64)
    in_channels = ckpt.get("config", {}).get("in_channels", 1024)
    print(f"[Phase 3] RSAdapter 설정: d_model={in_channels}, bottleneck={bottleneck}")

    # ── 4. Adapter 초기화 및 가중치 로드 ───────────────────────────────────
    adapters = nn.ModuleList([
        RSAdapter(d_model=in_channels, bottleneck=bottleneck)
        for _ in range(num_blocks)
    ]).to(device)

    adapters.load_state_dict(ckpt["adapters"])
    adapters.eval()
    print(f"[Phase 3] Adapter 가중치 로드 완료: {args.adapter_ckpt}")
    print(f"          (fpn / cls_head 가중치는 OVSS에서 사용하지 않음)")

    # ── 5. ViT 블록에 forward hook 등록 ───────────────────────────────────
    # 각 hook은 ViT block 출력을 받아 RSAdapter를 통과시키고 반환한다.
    # Sam3Processor.set_image() → backbone.forward_image() 호출 시 자동 적용됨.
    hooks = []
    for i in range(num_blocks):
        adapter = adapters[i]

        def _hook(_module, _input, output, _adapter=adapter):
            return _adapter(output)

        hooks.append(vit_blocks[i].register_forward_hook(_hook))

    print(f"[Phase 3] Forward hook {len(hooks)}개 등록 완료 (ViT 블록 0~{num_blocks-1})")

    # ── 6. OVSS 추론 실행 ──────────────────────────────────────────────────
    results = runner.test()

    # ── 7. Hook 해제 ───────────────────────────────────────────────────────
    for h in hooks:
        h.remove()
    print("[Phase 3] Forward hook 해제 완료")

    # ── 8. 결과 저장 ───────────────────────────────────────────────────────
    results.update({
        'Model': cfg.model.model_type + '+RSAdapter',
        'Dataset': cfg.dataset_type,
    })

    if runner.rank == 0:
        append_experiment_result('results.xlsx', [results])

    if runner.rank == 0:
        with open(os.path.join(cfg.work_dir, 'results.txt'), 'a') as f:
            f.write(osp.basename(args.config).split('.')[0] + '+adapter\n')
            for k, v in results.items():
                f.write(k + ': ' + str(v) + '\n')


if __name__ == '__main__':
    main()
